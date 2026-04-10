"""Tests for column_inference module."""

import datetime as dt

import pytest
from openpyxl import Workbook

from yumoyi_common.column_inference import (
    FieldSpec,
    SCORE_FORMAT_MATCH,
    SCORE_KEYWORD_CONTAINS,
    SCORE_KEYWORD_EXACT,
    EXCEL_SERIAL_MIN,
    EXCEL_SERIAL_MAX,
    _cell_to_str,
    _normalize_header,
    infer_columns,
    is_date_like,
    is_numeric,
)

from testdata.column_inference_zh import (
    HEADER_PRODUCT_NAME, HEADER_QTY, HEADER_DATE, HEADER_AMOUNT,
    HEADER_REMARK, HEADER_PRICE, HEADER_PRODUCT_CODE_LONG,
    HEADER_PRODUCT_CODE, HEADER_CODE, HEADER_COL_A, HEADER_COL_B,
    KW_PRODUCT_NAME, KW_QTY, KW_DATE, KW_AMOUNT, KW_CODE, KW_PRICE, KW_NAME,
    CELL_APPLE, CELL_BANANA, CELL_ORANGE, CELL_NON_NUMERIC_TEXT,
    NORMALIZE_INPUT, NORMALIZE_EXPECTED,
    FIELDSPEC_KEYWORDS,
)


# ============================================================
# _cell_to_str
# ============================================================

class TestCellToStr:
    def test_none(self):
        assert _cell_to_str(None) == ""

    def test_float_integer(self):
        assert _cell_to_str(3.0) == "3"

    def test_float_decimal(self):
        assert _cell_to_str(3.14) == "3.14"

    def test_bytes(self):
        assert _cell_to_str(b"hello") == "hello"

    def test_string(self):
        assert _cell_to_str("  abc  ") == "abc"

    def test_int(self):
        assert _cell_to_str(42) == "42"


# ============================================================
# _normalize_header
# ============================================================

class TestNormalizeHeader:
    def test_none(self):
        assert _normalize_header(None) == ""

    def test_chinese_with_spaces(self):
        assert _normalize_header(NORMALIZE_INPUT) == NORMALIZE_EXPECTED

    def test_english_case(self):
        assert _normalize_header("  Product Name  ") == "productname"

    def test_mixed_whitespace(self):
        assert _normalize_header("a\t b\n c") == "abc"


# ============================================================
# is_numeric
# ============================================================

class TestIsNumeric:
    def test_int(self):
        assert is_numeric(42) is True

    def test_float(self):
        assert is_numeric(3.14) is True

    def test_string_number(self):
        assert is_numeric("1,234.56") is True

    def test_non_number(self):
        assert is_numeric("hello") is False

    def test_empty_string(self):
        assert is_numeric("") is False

    def test_none(self):
        assert is_numeric(None) is False

    def test_bool_excluded(self):
        assert is_numeric(True) is False


# ============================================================
# is_date_like
# ============================================================

class TestIsDateLike:
    def test_datetime_object(self):
        assert is_date_like(dt.datetime(2024, 1, 15)) is True

    def test_date_object(self):
        assert is_date_like(dt.date(2024, 1, 15)) is True

    def test_date_string_dash(self):
        assert is_date_like("2024-01-15") is True

    def test_date_string_slash(self):
        assert is_date_like("2024/1/15") is True

    def test_excel_serial_number(self):
        assert is_date_like(45000) is True

    def test_non_date(self):
        assert is_date_like("hello") is False

    def test_number_outside_range(self):
        assert is_date_like(100) is False

    def test_none(self):
        assert is_date_like(None) is False

    def test_bool_excluded(self):
        assert is_date_like(True) is False
        assert is_date_like(False) is False

    def test_custom_serial_range(self):
        assert is_date_like(100) is False
        assert is_date_like(100, serial_range=(50, 200)) is True

    def test_constants_exported(self):
        assert EXCEL_SERIAL_MIN < EXCEL_SERIAL_MAX


# ============================================================
# FieldSpec
# ============================================================

class TestFieldSpec:
    def test_defaults(self):
        fs = FieldSpec("code")
        assert fs.name == "code"
        assert fs.required is False
        assert fs.keywords == ()
        assert fs.format_test is None
        assert fs.priority == 0

    def test_full_construction(self):
        fs = FieldSpec(
            "price",
            required=True,
            keywords=FIELDSPEC_KEYWORDS,
            format_test=is_numeric,
            priority=10,
        )
        assert fs.name == "price"
        assert fs.required is True
        assert fs.keywords == FIELDSPEC_KEYWORDS
        assert fs.format_test is is_numeric
        assert fs.priority == 10


# ============================================================
# infer_columns
# ============================================================

def _make_ws(headers, rows):
    """Helper: create an in-memory worksheet with given headers and data rows."""
    wb = Workbook()
    ws = wb.active
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    return ws


class TestInferColumns:
    def test_full_inference(self):
        """Headers + data -> correct mapping."""
        ws = _make_ws(
            [HEADER_PRODUCT_NAME, HEADER_QTY, HEADER_DATE],
            [
                [CELL_APPLE, 10, "2024-01-01"],
                [CELL_BANANA, 20, "2024-02-01"],
                [CELL_ORANGE, 30, "2024-03-01"],
            ],
        )
        specs = [
            FieldSpec("name", keywords=KW_PRODUCT_NAME, priority=1),
            FieldSpec("qty", keywords=KW_QTY, format_test=is_numeric, priority=0),
            FieldSpec("date", keywords=KW_DATE, format_test=is_date_like, priority=0),
        ]
        result = infer_columns(ws, specs)
        assert result["name"] == 1
        assert result["qty"] == 2
        assert result["date"] == 3

    def test_empty_worksheet(self):
        """Empty worksheet -> all None."""
        wb = Workbook()
        ws = wb.active
        specs = [FieldSpec("name", required=True, keywords=KW_NAME)]
        result = infer_columns(ws, specs)
        assert result["name"] is None

    def test_required_field_unmapped(self):
        """Required field not found -> value is None (caller decides what to do)."""
        ws = _make_ws([HEADER_COL_A, HEADER_COL_B], [["x", "y"]])
        specs = [FieldSpec("code", required=True, keywords=KW_CODE)]
        result = infer_columns(ws, specs)
        assert result["code"] is None

    def test_format_beats_keyword(self):
        """Format match score (100) > keyword exact score (80)."""
        # Col 1: header matches keyword but data is text
        # Col 2: header doesn't match but data is all numbers
        ws = _make_ws(
            [HEADER_AMOUNT, HEADER_REMARK],
            [
                [CELL_NON_NUMERIC_TEXT, 100],
                [CELL_NON_NUMERIC_TEXT, 200],
                [CELL_NON_NUMERIC_TEXT, 300],
            ],
        )
        specs = [
            FieldSpec(
                "amount",
                keywords=KW_AMOUNT,
                format_test=is_numeric,
                priority=0,
            ),
        ]
        result = infer_columns(ws, specs)
        assert result["amount"] == 2

    def test_no_duplicate_column_assignment(self):
        """Two fields competing for the same column -> higher priority wins."""
        ws = _make_ws(
            [HEADER_PRICE],
            [
                [99.9],
                [88.8],
                [77.7],
            ],
        )
        specs = [
            FieldSpec("cost", keywords=KW_PRICE, format_test=is_numeric, priority=10),
            FieldSpec("price", keywords=KW_PRICE, format_test=is_numeric, priority=0),
        ]
        result = infer_columns(ws, specs)
        assert result["cost"] == 1
        assert result["price"] is None

    def test_keyword_contains_match(self):
        """Partial keyword match still works."""
        ws = _make_ws([HEADER_PRODUCT_CODE_LONG], [["ABC123"]])
        specs = [FieldSpec("code", keywords=KW_CODE)]
        result = infer_columns(ws, specs)
        assert result["code"] == 1

    def test_keyword_exact_beats_contains(self):
        """Exact keyword match (80) beats contains match (50)."""
        ws = _make_ws([HEADER_PRODUCT_CODE, HEADER_CODE], [["ABC", "DEF"]])
        specs = [FieldSpec("code", keywords=KW_CODE)]
        result = infer_columns(ws, specs)
        # Col 2 header is exact match, col 1 only contains the keyword
        assert result["code"] == 2
